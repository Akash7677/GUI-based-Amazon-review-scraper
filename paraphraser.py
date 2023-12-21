
import time
import os
import openpyxl
import ai21
import pandas as pd

ai21.api_key = "iA8KLSIrCIPypWMzm0vSYzb2wf0GKLsT"

def paraphrased_text(text_to_process):
    max_retries = 5
    retry_delay = 60  # Initial delay in seconds
    retries = 0

    while retries < max_retries:
        try:
            if len(text_to_process) > 500:
                print("Skipped paraphrasing")
                return "--------------------Not Paraphrased---------------------------"
            response = ai21.Paraphrase.execute(
                text=f"{text_to_process}",
                style="casual"
            )
            paraphrased = response.suggestions[0].text
            return paraphrased
        except Exception as e:
            error_message = str(e)
            print(f"Error: {error_message}")

            # Check if the error message indicates a rate-limiting issue (HTTP 429)
            if "429" in error_message:
                print(f"Received 429 response. Retrying in {retry_delay} seconds...")
                time.sleep(retry_delay)
                retries += 1
                retry_delay *= 2  # Exponential backoff
            else:
                print(f"Retrying in {retry_delay} seconds...")
                time.sleep(retry_delay)
                retries += 1
                retry_delay *= 2  # Exponential backoff

        raise Exception(f"Failed to get a valid response after {max_retries} retries.")


def process_excel_file(file_path):
    # global count
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(file_path)
    print(f"Current files: {file_path}")
    # Assuming you are working with the first sheet, change accordingly if needed
    sheet = workbook.active

    sheet.cell(row=1, column=4).value = "Paraphrased"
    # Iterate through column C starting from index 2
    last_row = sheet.max_row
    rev_done = 0
    for row_index in range(2, last_row+1):
        try:
            # Get the value in column C
            # print(f"Reading text from C{row_index}")
            original_text = sheet.cell(row=row_index, column=3).value
            # count += 1
            # if count == 29:
            #     # API cooldown period
            #     print("~"*80)
            #     print("Paraphrase API Cooldown of 60s")
            #     print("~" * 80)
            #     time.sleep(60)
            #     count = 0
            # Pass the content of column C to the paraphrased_text function
            paraphrased = paraphrased_text(original_text)
            rev_done += 1
            # print(f"writing to D{row_index}")
            # Save the output of the function in column D
            sheet.cell(row=row_index, column=4).value = paraphrased
            print(f"Reviews paraphrased: {rev_done}")
        except Exception as e:
            # Save the changes to the Excel file
            workbook.save(file_path)
            print(f"Saved.....due to {e}")
            continue
    workbook.save(file_path)
    print("Saved.....")



def extract_p_revs(filepath):
    final_revs = []
    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook.active
    last_row = sheet.max_row
    count = 0
    for row_index in range(2, last_row + 1):
        # Get the value in column D
        title = sheet.cell(row=row_index, column=2).value
        reviews = sheet.cell(row=row_index, column=4).value
        # Check if the value is not None
        if reviews is not None:
            # Check if "Not Paraphrased" is in the reviews
            if "Not Paraphrased" in reviews:
                continue
            data = {
                "Title": title,
                "Reviews": reviews
            }
            final_revs.append(data)

    df = pd.DataFrame(final_revs)
    df = df.sample(frac=1).reset_index(drop=True)
    print(f"Extracted {len(df)} reviews")
    return df

def merge_revs(folder_path):
    Final_revs_file = os.path.join(folder_path, "final_reviews.xlsx")
    files = os.listdir(folder_path)
    # Filter out only Excel files
    excel_files = [file for file in files if file.endswith('.xlsx')]
    if not excel_files:
        print("No Excel files found in the specified folder.")
        return
    print(f"Found {len(excel_files)} Excel file(s) in the folder.")
    final_df = pd.DataFrame()
    for excel_file in excel_files:
        excel_file = os.path.join(folder_path, excel_file)
        rev_df = extract_p_revs(excel_file)
        final_df = pd.concat([final_df, rev_df], ignore_index=True)

    if os.path.exists(Final_revs_file):
        existing_df = pd.read_excel(Final_revs_file, engine='openpyxl')
        updated_df = pd.concat([existing_df, final_df], ignore_index=True)
        updated_df.to_excel(Final_revs_file, index=False, engine='openpyxl')
    else:
        final_df.to_excel(Final_revs_file, index=False, engine='openpyxl')


# Specify the path to your Excel file
def main_para(folder_path=str(os.getcwd())):
    # Get a list of all files in the specified folder
    files = os.listdir(folder_path)
    # Filter out only Excel files
    excel_files = [file for file in files if file.endswith('.xlsx')]
    if not excel_files:
        print("No Excel files found in the specified folder.")
        return
    print(f"Found {len(excel_files)} Excel file(s) in the folder.")
    # Process each Excel file
    for excel_file in excel_files:
        # Build the full path to the Excel file
        excel_file_path = os.path.join(folder_path, excel_file)
        # Call the function to process the Excel file
        process_excel_file(excel_file_path)
    merge_revs(folder_path)

if __name__ == "__main__":
    merge_revs(folder_path='/home/vvdn/pythonProject/scrapper_ui/test')